from __future__ import annotations

from pathlib import Path
import csv

from openpyxl import load_workbook


HEADER_ALIASES = {
    "display_name": {"name", "display_name", "traveler", "이름", "성명", "출장자"},
    "english_name": {"english_name", "english", "eng_name", "영문명", "영어이름"},
    "aliases": {"aliases", "alias", "other_names", "별칭", "다른이름"},
    "affiliation": {"affiliation", "department", "lab", "소속", "부서", "연구실"},
}


def normalize_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def field_for_header(header: str) -> str | None:
    header = normalize_header(header)
    for field, aliases in HEADER_ALIASES.items():
        if header in aliases:
            return field
    return None


def load_roster(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return load_xlsx_roster(path)
    return load_csv_roster(path)


def load_csv_roster(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows = list(reader)
    return rows_to_records(rows)


def load_xlsx_roster(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
    return rows_to_records(rows)


def rows_to_records(rows: list[list[object]]) -> list[dict[str, str]]:
    rows = [row for row in rows if any(str(cell or "").strip() for cell in row)]
    if not rows:
        return []
    headers = [field_for_header(str(cell)) for cell in rows[0]]
    records: list[dict[str, str]] = []
    for row in rows[1:]:
        record = {"display_name": "", "english_name": "", "aliases": "", "affiliation": ""}
        for idx, cell in enumerate(row):
            if idx >= len(headers) or headers[idx] is None:
                continue
            record[headers[idx] or ""] = str(cell or "").strip()
        if record["display_name"]:
            records.append(record)
    return records

