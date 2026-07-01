from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import shutil
import zipfile

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas


FONT_NAME = "Helvetica"


def register_font() -> str:
    global FONT_NAME
    candidates = [
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for candidate in candidates:
        path = Path(candidate)
        if path.exists():
            try:
                pdfmetrics.registerFont(TTFont("TripDocFont", str(path)))
                FONT_NAME = "TripDocFont"
                return FONT_NAME
            except Exception:
                continue
    return FONT_NAME


def excel_summary(path: Path, trip: dict, travelers: list[dict], documents: list[dict]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Documents"
    ws.append(
        [
            "document_id",
            "traveler",
            "category",
            "status",
            "amount_hint",
            "date_hint",
            "uploaded_by",
            "filename",
            "notes",
        ]
    )
    for doc in documents:
        ws.append(
            [
                doc["id"],
                doc.get("traveler_name") or "",
                doc["category"],
                doc["status"],
                doc["amount_hint"],
                doc["date_hint"],
                doc["uploaded_by"],
                doc["original_filename"],
                doc["notes"],
            ]
        )

    ws2 = wb.create_sheet("Travelers")
    ws2.append(["traveler_id", "name", "english_name", "aliases", "affiliation"])
    for traveler in travelers:
        ws2.append(
            [
                traveler["id"],
                traveler["display_name"],
                traveler["english_name"],
                traveler["aliases"],
                traveler["affiliation"],
            ]
        )

    ws3 = wb.create_sheet("Trip")
    ws3.append(["field", "value"])
    ws3.append(["trip_id", trip["id"]])
    ws3.append(["name", trip["name"]])
    ws3.append(["description", trip["description"]])
    wb.save(path)
    return path


def safe_pdf_text(value: object) -> str:
    text = str(value or "")
    if FONT_NAME == "Helvetica":
        return text.encode("latin-1", errors="replace").decode("latin-1")
    return text


def person_pdf(path: Path, trip: dict, traveler: dict, documents: list[dict]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    register_font()
    pdf = canvas.Canvas(str(path), pagesize=A4)
    width, height = A4
    y = height - 48
    pdf.setFont(FONT_NAME, 14)
    pdf.drawString(48, y, safe_pdf_text(f"Trip document summary: {trip['name']}"))
    y -= 26
    pdf.setFont(FONT_NAME, 11)
    pdf.drawString(48, y, safe_pdf_text(f"Traveler: {traveler['display_name']}"))
    y -= 28
    pdf.setFont(FONT_NAME, 9)
    for doc in documents:
        lines = [
            f"File: {doc['original_filename']}",
            f"Category: {doc['category']} | Status: {doc['status']}",
            f"Date: {doc['date_hint']} | Amount: {doc['amount_hint']}",
            f"Notes: {doc['notes']}",
        ]
        for line in lines:
            if y < 60:
                pdf.showPage()
                pdf.setFont(FONT_NAME, 9)
                y = height - 48
            pdf.drawString(48, y, safe_pdf_text(line[:120]))
            y -= 14
        y -= 8
    pdf.save()
    return path


def zip_export(
    zip_path: Path,
    trip: dict,
    travelers: list[dict],
    documents: list[dict],
    root_dir: Path,
) -> Path:
    export_dir = zip_path.parent / "package"
    if export_dir.exists():
        shutil.rmtree(export_dir)
    export_dir.mkdir(parents=True, exist_ok=True)

    excel_path = excel_summary(export_dir / "summary.xlsx", trip, travelers, documents)
    by_traveler: dict[int, list[dict]] = defaultdict(list)
    for doc in documents:
        if doc["traveler_id"]:
            by_traveler[int(doc["traveler_id"])].append(doc)

    traveler_lookup = {int(item["id"]): item for item in travelers}
    pdf_dir = export_dir / "person-pdfs"
    for traveler_id, docs in by_traveler.items():
        traveler = traveler_lookup.get(traveler_id)
        if traveler:
            filename = f"{traveler['display_name']}_{traveler_id}.pdf"
            person_pdf(pdf_dir / filename, trip, traveler, docs)

    originals_dir = export_dir / "originals"
    originals_dir.mkdir(exist_ok=True)
    for doc in documents:
        source = root_dir / doc["stored_path"]
        if source.exists():
            target_name = f"{doc['id']:04d}_{doc['original_filename']}"
            shutil.copy2(source, originals_dir / target_name)

    zip_path.parent.mkdir(parents=True, exist_ok=True)
    if zip_path.exists():
        zip_path.unlink()
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for file in export_dir.rglob("*"):
            if file.is_file():
                zf.write(file, file.relative_to(export_dir))
    return zip_path

